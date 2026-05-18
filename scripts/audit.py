#!/usr/bin/env python3
"""Audit: Supabase vs Excel data comparison for PetraLabs RRHH."""
import json
import sys
import urllib.request
import urllib.error
from pathlib import Path
from collections import OrderedDict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip3 install openpyxl")
    sys.exit(1)

# ============================================================
# CONFIG
# ============================================================
SUPABASE_URL = "https://zgzsggetnlbibjembesc.supabase.co"
ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InpnenNnZ2V0bmxiaWJqZW1iZXNjIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzg0NDg5MTIsImV4cCI6MjA5NDAyNDkxMn0.77qhtGIx16xiQu3P-Vti4L6JJ4MisgpK4OHmC6QMFSs"
DATA_RAW = Path("/home/nch/Escritorio/Things/RRHH/data-raw")

TABLES = [
    "empleados",
    "home_office_semanal",
    "vacaciones",
    "vacaciones_dias",
    "lineas_moviles",
    "sueldos",
    "manuales",
    "eventos_calendario",
]

EXCEL_FILES = [
    "Legajo Colaboradores.xlsx",
    "\U0001f4b0Sueldos Contexto.xlsx",  # moneybag emoji
    "Listado de Manuales por Area.xlsx",
]

# ============================================================
# HELPERS
# ============================================================
def query_supabase(table: str, select: str = "*") -> dict:
    """Query Supabase REST API. Returns dict with 'data', 'count', 'error'."""
    url = f"{SUPABASE_URL}/rest/v1/{table}?select={select}"
    headers = {
        "apikey": ANON_KEY,
        "Authorization": f"Bearer {ANON_KEY}",
        "Accept": "application/json",
        "Prefer": "count=exact",
    }
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
            count = resp.headers.get("content-range", "0-0/0").split("/")[-1]
            return {"data": data, "count": int(count), "error": None}
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        return {"data": None, "count": 0, "error": f"HTTP {e.code}: {body[:200]}"}
    except Exception as e:
        return {"data": None, "count": 0, "error": str(e)[:200]}


def query_supabase_raw(table: str, select: str = "*") -> dict:
    """Query Supabase with raw SQL-like select (no count)."""
    url = f"{SUPABASE_URL}/rest/v1/{table}?select={select}"
    headers = {
        "apikey": ANON_KEY,
        "Authorization": f"Bearer {ANON_KEY}",
        "Accept": "application/json",
    }
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
            return {"data": data, "error": None}
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        return {"data": None, "error": f"HTTP {e.code}: {body[:200]}"}
    except Exception as e:
        return {"data": None, "error": str(e)[:200]}


def excel_sheet_names(path: Path):
    """Return list of sheet names in an Excel file."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return wb.sheetnames
    except Exception as e:
        return [f"ERROR: {e}"]


def excel_read_sheet(path: Path, sheet: str):
    """Read a sheet into a list of dicts (header -> rows)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            row_dict = OrderedDict()
            for i, val in enumerate(row):
                if i < len(header):
                    row_dict[header[i]] = val
            result.append(row_dict)
        return result
    except KeyError:
        return [f"ERROR: Sheet '{sheet}' not found"]
    except Exception as e:
        return [f"ERROR: {e}"]


def analyze_null_cols(rows: list, label: str):
    """Report columns where ALL values are None."""
    if not rows or not isinstance(rows[0], dict):
        return []
    cols = list(rows[0].keys())
    null_cols = []
    for col in cols:
        all_null = all(row.get(col) is None for row in rows)
        if all_null:
            null_cols.append(col)
    return null_cols


def col_set(rows: list):
    """Return set of column names from list of dicts."""
    if not rows or not isinstance(rows[0], dict):
        return set()
    return set(rows[0].keys())


def fmt_val(v):
    """Format a value for display."""
    if v is None:
        return "NULL"
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v)


# ============================================================
# MAIN
# ============================================================
def main():
    report = []
    report.append("# ============================================================")
    report.append("# AUDIT REPORT: Supabase DB vs Excel Data")
    report.append(f"# Generated: {datetime.now().isoformat()}")
    report.append("# ============================================================\n")

    # ---- PART 1: SUPABASE DATA ----
    report.append("## PART 1: SUPABASE DATABASE STATE\n")

    db_data = {}
    for table in TABLES:
        print(f"  Querying Supabase: {table}...")
        result = query_supabase(table)
        db_data[table] = result

        if result["error"]:
            if "406" in str(result["error"]) or "416" in str(result["error"]):
                # Might be too many rows, try without count
                r2 = query_supabase_raw(table)
                if r2["data"] is not None:
                    result["data"] = r2["data"]
                    result["count"] = len(r2["data"])
                    result["error"] = None
                else:
                    report.append(f"### {table}")
                    report.append(f"  ERROR: {result['error']}\n")
                    continue
            else:
                report.append(f"### {table}")
                report.append(f"  ERROR: {result['error']}\n")
                continue

        report.append(f"### {table}")
        report.append(f"  Row count: {result['count']}")

        if result["data"] and len(result["data"]) > 0:
            sample = result["data"][0]
            report.append(f"  Columns ({len(sample)}): {list(sample.keys())}")

            # Check for all-null columns
            null_cols = analyze_null_cols(result["data"], table)
            if null_cols:
                report.append(f"  ⚠️  ALL-NULL COLUMNS: {null_cols}")
            else:
                report.append(f"  No all-null columns found.")

            # Show first 2 rows as sample
            report.append(f"  Sample rows:")
            for i, row in enumerate(result["data"][:2]):
                for k, v in row.items():
                    report.append(f"    {k}: {fmt_val(v)}")
                if i < 1:
                    report.append(f"    ---")
        else:
            report.append(f"  No data returned (empty table or RLS blocked).")
        report.append("")

    # ---- PART 2: EXCEL DATA ----
    report.append("\n## PART 2: EXCEL RAW DATA\n")

    excel_data = {}
    for fname in EXCEL_FILES:
        fpath = DATA_RAW / fname
        if not fpath.exists():
            # try with emoji
            for p in DATA_RAW.iterdir():
                if p.name.endswith("Sueldos Contexto.xlsx") or fname.replace("\U0001f4b0", "") in p.name:
                    fpath = p
                    break
        print(f"  Reading Excel: {fpath.name}...")
        sheets = excel_sheet_names(fpath)
        report.append(f"### 📄 {fpath.name}")
        report.append(f"  Sheets: {sheets}")
        for s_name in sheets:
            sheet_data = excel_read_sheet(fpath, s_name)
            excel_data[(fpath.name, s_name)] = sheet_data
            if not sheet_data:
                report.append(f"  Sheet '{s_name}': EMPTY")
                continue
            if isinstance(sheet_data[0], str) and sheet_data[0].startswith("ERROR"):
                report.append(f"  Sheet '{s_name}': {sheet_data[0]}")
                continue
            report.append(f"\n  Sheet: **{s_name}**")
            report.append(f"    Rows: {len(sheet_data)}")
            report.append(f"    Columns: {list(sheet_data[0].keys())}")
            # Null columns in excel
            null_cols = analyze_null_cols(sheet_data, f"{fpath.name}/{s_name}")
            if null_cols:
                report.append(f"    ⚠️  ALL-NULL COLUMNS in Excel: {null_cols}")
            # Show first 2 rows
            report.append(f"    Sample rows:")
            for i, row in enumerate(sheet_data[:2]):
                for k, v in row.items():
                    report.append(f"      {k}: {fmt_val(v)}")
                if i < 1:
                    report.append(f"      ---")
            report.append("")

    # ---- PART 3: COMPARISON ----
    report.append("\n## PART 3: COMPARISON: Excel vs Database\n")

    # 3a. Legajo sheet -> empleados
    print("  Comparing Legajo -> empleados...")
    legajo_data = excel_data.get(("Legajo Colaboradores.xlsx", "Legajo"), [])
    if legajo_data and isinstance(legajo_data[0], dict):
        db_emp = db_data.get("empleados", {}).get("data", [])
        excel_cols = col_set(legajo_data)
        db_cols = col_set(db_emp) if db_emp else set()

        report.append("### 3a. Legajo (Excel) → empleados (DB)\n")
        report.append(f"  Excel columns ({len(excel_cols)}): {sorted(excel_cols)}")
        report.append(f"  DB columns ({len(db_cols)}): {sorted(db_cols)}")

        in_excel_not_db = excel_cols - db_cols
        in_db_not_excel = db_cols - excel_cols
        if in_excel_not_db:
            report.append(f"  ❌ IN EXCEL BUT NOT IN DB: {sorted(in_excel_not_db)}")
        else:
            report.append(f"  ✅ All Excel columns exist in DB.")
        if in_db_not_excel:
            # Filter out metadata columns
            meta_cols = {c for c in in_db_not_excel if c in ("id", "empresa_id", "created_at", "updated_at")}
            real_cols = in_db_not_excel - meta_cols
            if real_cols:
                report.append(f"  ℹ️  IN DB BUT NOT EXCEL: {sorted(real_cols)} (expected metadata: {sorted(meta_cols)})")
            else:
                report.append(f"  ℹ️  Only metadata cols in DB not in Excel: {sorted(meta_cols)}")

        # Check for all-null in DB
        if db_emp:
            null_cols = analyze_null_cols(db_emp, "empleados")
            if null_cols:
                report.append(f"  ⚠️  DB NULLS: columns fully NULL: {null_cols}")

        # Row count comparison
        report.append(f"\n  Row count: Excel={len(legajo_data)}, DB={len(db_emp)}")
        if len(legajo_data) != len(db_emp):
            report.append(f"  ⚠️  MISMATCH: {'Excel has more rows' if len(legajo_data) > len(db_emp) else 'DB has more rows'} by {abs(len(legajo_data)-len(db_emp))}")
        report.append("")

    # 3b. HO sheet -> home_office_semanal
    print("  Comparing HO -> home_office_semanal...")
    ho_data = excel_data.get(("Legajo Colaboradores.xlsx", "HO"), [])
    if ho_data and isinstance(ho_data[0], dict):
        db_ho = db_data.get("home_office_semanal", {}).get("data", [])
        excel_cols = col_set(ho_data)
        db_cols = col_set(db_ho) if db_ho else set()
        report.append("### 3b. HO (Excel) → home_office_semanal (DB)\n")
        report.append(f"  Excel columns ({len(excel_cols)}): {sorted(excel_cols)}")
        report.append(f"  DB columns ({len(db_cols)}): {sorted(db_cols)}")
        in_excel_not_db = excel_cols - db_cols
        in_db_not_excel = db_cols - excel_cols
        if in_excel_not_db:
            report.append(f"  ❌ IN EXCEL BUT NOT IN DB: {sorted(in_excel_not_db)}")
        else:
            report.append(f"  ✅ All Excel columns exist in DB.")
        if in_db_not_excel:
            meta_cols = {c for c in in_db_not_excel if c in ("id", "created_at", "unique")}
            real_cols = in_db_not_excel - meta_cols
            if real_cols or True:
                report.append(f"  IN DB BUT NOT EXCEL: {sorted(in_db_not_excel)}")
        report.append(f"\n  Row count: Excel={len(ho_data)}, DB={len(db_ho)}")
        if len(ho_data) != len(db_ho):
            report.append(f"  ⚠️  MISMATCH: DB has {len(db_ho)}, Excel has {len(ho_data)}")
        report.append("")

    # 3c. Vacaciones 2025 sheet -> vacaciones
    print("  Comparing Vacaciones 2025 -> vacaciones...")
    vac_data = excel_data.get(("Legajo Colaboradores.xlsx", "Vacaciones 2025"), [])
    if vac_data and isinstance(vac_data[0], dict):
        db_vac = db_data.get("vacaciones", {}).get("data", [])
        db_vac_dias = db_data.get("vacaciones_dias", {}).get("data", [])
        excel_cols = col_set(vac_data)
        db_cols = col_set(db_vac) if db_vac else set()
        report.append("### 3c. Vacaciones 2025 (Excel) → vacaciones + vacaciones_dias (DB)\n")
        report.append(f"  Excel columns ({len(excel_cols)}): {sorted(excel_cols)}")
        report.append(f"  DB vacaciones columns ({len(db_cols)}): {sorted(db_cols)}")
        in_excel_not_db = excel_cols - db_cols - {"Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"}
        # The month columns in Excel likely correspond to vacaciones_dias
        month_cols = excel_cols & {"Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"}
        if month_cols:
            report.append(f"  ℹ️  Month columns in Excel (→ vacaciones_dias.mes): {sorted(month_cols)}")
        if in_excel_not_db:
            report.append(f"  ❌ IN EXCEL BUT NOT IN DB: {sorted(in_excel_not_db)}")
        report.append(f"\n  Row count: Excel={len(vac_data)}, DB vacaciones={len(db_vac)}, DB vacaciones_dias={len(db_vac_dias)}")
        report.append("")

    # 3d. Lineas Móviles sheet -> lineas_moviles
    print("  Comparing Lineas Móviles -> lineas_moviles...")
    lm_data = excel_data.get(("Legajo Colaboradores.xlsx", "Lineas Móviles"), [])
    if lm_data and isinstance(lm_data[0], dict):
        db_lm = db_data.get("lineas_moviles", {}).get("data", [])
        excel_cols = col_set(lm_data)
        db_cols = col_set(db_lm) if db_lm else set()
        report.append("### 3d. Lineas Móviles (Excel) → lineas_moviles (DB)\n")
        report.append(f"  Excel columns ({len(excel_cols)}): {sorted(excel_cols)}")
        report.append(f"  DB columns ({len(db_cols)}): {sorted(db_cols)}")
        in_excel_not_db = excel_cols - db_cols
        if in_excel_not_db:
            report.append(f"  ❌ IN EXCEL BUT NOT IN DB: {sorted(in_excel_not_db)}")
        else:
            report.append(f"  ✅ All Excel columns exist in DB.")
        in_db_not_excel = db_cols - excel_cols
        if in_db_not_excel:
            report.append(f"  IN DB BUT NOT EXCEL: {sorted(in_db_not_excel)}")
        report.append(f"\n  Row count: Excel={len(lm_data)}, DB={len(db_lm)}")
        if len(lm_data) != len(db_lm):
            report.append(f"  ⚠️  MISMATCH: {'Excel has more' if len(lm_data) > len(db_lm) else 'DB has more'} rows")
        report.append("")

    # 3e. Sueldos Contexto -> sueldos
    print("  Comparing Sueldos Contexto -> sueldos...")
    sueldos_fname = None
    for k in excel_data:
        if "Sueldos" in k[0]:
            sueldos_fname = k[0]
    for sname_key in [k for k in excel_data if sueldos_fname and k[0] == sueldos_fname]:
        s_name = sname_key[1]
        s_data = excel_data[sname_key]
        if not s_data or not isinstance(s_data[0], dict):
            continue
        db_sue = db_data.get("sueldos", {}).get("data", [])
        excel_cols = col_set(s_data)
        db_cols = col_set(db_sue) if db_sue else set()
        report.append(f"### 3e. 💰Sueldos Contexto.xlsx → Sheet '{s_name}' → sueldos (DB)\n")
        report.append(f"  Excel '{s_name}' columns ({len(excel_cols)}): {sorted(excel_cols)}")
        report.append(f"  DB sueldos columns ({len(db_cols)}): {sorted(db_cols)}")
        in_excel_not_db = excel_cols - db_cols
        if in_excel_not_db:
            report.append(f"  ❌ IN EXCEL BUT NOT IN DB: {sorted(in_excel_not_db)}")
        in_db_not_excel = db_cols - excel_cols
        if in_db_not_excel:
            report.append(f"  IN DB BUT NOT EXCEL: {sorted(in_db_not_excel)}")
        report.append(f"\n  Row count: Excel='{s_name}'={len(s_data)}, DB sueldos={len(db_sue)}")

        # Show a sample row
        report.append(f"  Sample row (Excel):")
        for k, v in list(s_data[0].items())[:10]:
            report.append(f"    {k}: {fmt_val(v)}")
        report.append("")

    # 3f. Manuales -> manuales
    print("  Comparing Manuales -> manuales...")
    man_data = excel_data.get(("Listado de Manuales por Area.xlsx", None))
    # find the correct key
    for k in excel_data:
        if "Manuales" in k[0]:
            man_data = excel_data[k]
            man_sheet = k[1]
            break
    if man_data and isinstance(man_data[0], dict):
        db_man = db_data.get("manuales", {}).get("data", [])
        excel_cols = col_set(man_data)
        db_cols = col_set(db_man) if db_man else set()
        report.append(f"### 3f. Listado de Manuales por Area.xlsx → Sheet '{man_sheet}' → manuales (DB)\n")
        report.append(f"  Excel columns ({len(excel_cols)}): {sorted(excel_cols)}")
        report.append(f"  DB columns ({len(db_cols)}): {sorted(db_cols)}")
        in_excel_not_db = excel_cols - db_cols
        if in_excel_not_db:
            report.append(f"  ❌ IN EXCEL BUT NOT IN DB: {sorted(in_excel_not_db)}")
        in_db_not_excel = db_cols - excel_cols
        if in_db_not_excel:
            report.append(f"  IN DB BUT NOT EXCEL: {sorted(in_db_not_excel)}")
        report.append(f"\n  Row count: Excel={len(man_data)}, DB={len(db_man)}")
        if len(man_data) != len(db_man):
            report.append(f"  ⚠️  MISMATCH: {'Excel has more' if len(man_data) > len(db_man) else 'DB has more'} rows")
        report.append("")

    # 3g. Sheets without DB tables
    print("  Checking for sheets without DB tables...")
    all_sheets = set()
    for fname in EXCEL_FILES:
        fpath = DATA_RAW / fname
        if not fpath.exists():
            for p in DATA_RAW.iterdir():
                if fname.replace("\U0001f4b0", "") in p.name:
                    fpath = p
                    break
        sheets = excel_sheet_names(fpath)
        for s in sheets:
            all_sheets.add((fpath.name, s))

    mapped_sheets = {
        ("Legajo Colaboradores.xlsx", "Legajo"): "empleados",
        ("Legajo Colaboradores.xlsx", "HO"): "home_office_semanal",
        ("Legajo Colaboradores.xlsx", "Vacaciones 2025"): "vacaciones + vacaciones_dias",
        ("Legajo Colaboradores.xlsx", "Lineas Móviles"): "lineas_moviles",
        ("💰Sueldos Contexto.xlsx", None): "sueldos",
        ("Listado de Manuales por Area.xlsx", None): "manuales",
    }

    report.append("### 3g. Sheets in Excel with NO corresponding DB table\n")
    for fname, sname in sorted(all_sheets):
        is_mapped = False
        for (ef, es), tbl in mapped_sheets.items():
            if ef == fname and (es is None or es == sname):
                is_mapped = True
                break
        if not is_mapped:
            report.append(f"  ❌ '{fname}' → Sheet '{sname}' has NO DB table")
    # Also check reuniones
    report.append("")
    report.append("#### Reuniones (frontend only)\n")
    report.append("  The 'Reuniones' feature exists only in the frontend (src/features/reuniones/)")
    report.append("  with columns: id, titulo, fecha, hora, duracion, participantes, resumen")
    report.append("  and returns empty data. There is NO Supabase table for reuniones.")
    report.append("  ")

    # ---- PART 4: FORMAT DIFFERENCES ----
    report.append("\n## PART 4: FORMAT DIFFERENCES\n")

    # Legajo: Excel columns mapping to DB
    if legajo_data and isinstance(legajo_data[0], dict):
        report.append("### 4a. Legajo Colaboradores → empleados\n")
        legajo_row = legajo_data[0]
        # Compare types
        for col, val in legajo_row.items():
            db_col_map = {
                "Nombre y Apellido": "nombre_apellido",
                "Fecha Nacimiento": "fecha_nacimiento",
                "Documento": "dni",
                "Celular": "celular",
                "Contacto de Emergencia": "contacto_emergencia",
                "Fecha de Ingreso": "fecha_ingreso",
                "Email": "email",
                "Dirección": "direccion",
                "Movilidad": "movilidad",
                "Equipo": "equipo_ingreso",
            }
            if col in db_col_map:
                report.append(f"  '{col}' → '{db_col_map[col]}'")
                if isinstance(val, datetime):
                    report.append(f"    Value type: datetime (DB expects date)")
            elif col not in ("Activo",):
                report.append(f"  '{col}' → (no direct mapping in DB)")

    report.append("\n---\n")
    report.append("# END OF REPORT")

    # Write report
    output = "\n".join(report)
    outpath = Path("/home/nch/Escritorio/Things/RRHH/audit_report.md")
    outpath.write_text(output)
    print(f"\n✅ Report written to: {outpath}")
    print(f"   Total lines: {len(report)}")

    # Also print summary
    print("\n=== QUICK SUMMARY ===")
    for table in TABLES:
        d = db_data.get(table, {})
        if d.get("error"):
            print(f"  {table}: ERROR - {d['error'][:60]}")
        elif d.get("data") is not None:
            print(f"  {table}: {d['count']} rows")
        else:
            print(f"  {table}: no data")


if __name__ == "__main__":
    main()
