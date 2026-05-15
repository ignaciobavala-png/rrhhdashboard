"""
Genera SQL INSERT para importar días de estudio y mudanza desde el Excel
a la tabla eventos_calendario.

Uso:
    python scripts/import-eventos-calendario.py

Genera un bloque SQL listo para pegar en el SQL Editor de Supabase.
"""

import openpyxl
import re

EXCEL_PATH = "data-raw/Legajo Colaboradores.xlsx"

MESES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4,
    "may": 5, "jun": 6, "jul": 7, "ago": 8,
    "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}


def find_empleado_id(nombre_apellido: str) -> int | None:
    """Hardcodeo del mapping nombre → id (consultado de la DB)."""
    mapping = {
        "HERNANDEZ, IGNACIO": 17,
        "CIRAVEGNA, FRANCO": 3,
        "CANOSA, LARA": 9,
        "REINERO, NICOLAS": 7,
    }
    return mapping.get(nombre_apellido.upper().strip())


def parse_study_dates(raw_dates: list[str]) -> list[tuple[str, str | None]]:
    """Retorna lista de (fecha YYYY-MM-DD, descripcion opcional)."""
    results: list[tuple[str, str | None]] = []
    for raw in raw_dates:
        raw = raw.strip()
        if not raw:
            continue

        # "22 y 24 sept"
        if " y " in raw:
            parts = re.split(r"\s+y\s+", raw)
            mes_encontrado = None
            for p in parts:
                for mname in MESES:
                    if mname in p.lower():
                        mes_encontrado = MESES[mname]
                        break
            if mes_encontrado:
                for p in parts:
                    day = "".join(c for c in p if c.isdigit())
                    if day:
                        results.append((f"2025-{mes_encontrado:02d}-{int(day):02d}", None))
            continue

        # "6 de Feb"
        m = re.match(r"(\d+)\s+de\s+(\w+)", raw, re.IGNORECASE)
        if m:
            day = int(m.group(1))
            mes_text = m.group(2).lower()[:3]
            if mes_text in MESES:
                results.append((f"2025-{MESES[mes_text]:02d}-{day:02d}", None))
            continue

        # "14 de Feb" (sin "de")
        parts = raw.split()
        if len(parts) >= 2:
            day = "".join(c for c in parts[0] if c.isdigit())
            mes_text = parts[-1].lower()[:3]
            if day and mes_text in MESES:
                results.append((f"2025-{MESES[mes_text]:02d}-{int(day):02d}", None))

    return results


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Vacaciones 2025"]

    inserts: list[str] = []

    # ── Días de estudio: Nacho (Ignacio Hernandez) ──
    row21 = ws[21]
    emp_id = find_empleado_id("HERNANDEZ, Ignacio")
    assert emp_id, "No se encontró ID para Hernandez"
    raw_dates = [str(c.value) for c in row21[1:6] if c.value]
    if row21[15].value:
        raw_dates.append(str(row21[15].value))
    parsed = parse_study_dates(raw_dates)
    for fecha, desc in parsed:
        inserts.append(f"  ({emp_id}, 'estudio', '{fecha}', null),")

    # ── Días de estudio: Franco Ciravegna ──
    row22 = ws[22]
    emp_id = find_empleado_id("CIRAVEGNA, Franco")
    assert emp_id
    raw_dates = [str(c.value) for c in row22[1:6] if c.value]
    parsed = parse_study_dates(raw_dates)
    for fecha, desc in parsed:
        inserts.append(f"  ({emp_id}, 'estudio', '{fecha}', null),")

    # ── Saldos de estudio (solo registro informativo) ──
    ws2 = wb["Vacaciones"]
    saldos_info = [
        (23, "IH", "HERNANDEZ, Ignacio"),
        (24, "LC", "CANOSA, Lara"),
        (25, "NR", "REINERO, Nicolas"),
    ]
    for row_idx, inicial, full_name in saldos_info:
        row = ws2[row_idx]
        saldo = row[15].value if len(row) > 15 else None  # Columna O = índice 15 (0-based)
        print(f"  -- {full_name} (inicial {inicial}): saldo estudio = {saldo}")

    # ── Generar SQL ──
    print("-- ============================================================")
    print("-- Importación de días de estudio desde Excel")
    print("-- Generado por scripts/import-eventos-calendario.py")
    print("-- ============================================================")
    print()
    if inserts:
        print("insert into eventos_calendario (empleado_id, tipo, fecha, descripcion)")
        print("values")
        for i, line in enumerate(inserts):
            sep = "," if i < len(inserts) - 1 else ";"
            print(f"{line.rstrip(',')}{sep}")
        print()
    else:
        print("-- No se encontraron días de estudio para importar.")
        print()

    print("-- Para ejecutar, pegá este SQL en el SQL Editor de Supabase.")
    print(f"-- Total registros: {len(inserts)}")


if __name__ == "__main__":
    main()
